import re
import pandas as pd
import openpyxl
import glob
import msvcrt
import os


# File preprocessing: Remove irrelevant information at the front end.
def dealtxt(infile):
    input_file = infile
    output_file = 'deal.txt'
    target_field = "# Detailed group assignments are listed below"

    with open(input_file, 'r') as f_in:
        lines = f_in.readlines()

    # Find the index of the row containing the target field.
    start_index = None
    for i, line in enumerate(lines):
        if target_field in line:
            start_index = i
            break

    # Remove the invalid content before the target content.
    if start_index is not None:
        new_content = ''.join(lines[start_index + 1:])

        # Write the new content and output it to a file.
        with open(output_file, 'w') as f_out:
            f_out.write(new_content)


# Open the file with the filename passed in as "i".
def open_file(i):
    filenames = i
    with open(filenames) as f:
        files = f.read()
    return files


# Generate Excel file according to the dictionary.
def createxl(filenames,name):
    # Create a DataFrame.
    df = pd.DataFrame.from_dict(filenames, orient='index')

    # Create an Excel Writer object.
    excel_writer = pd.ExcelWriter(name+'.xlsx', engine='openpyxl')

    # Write the DataFrame to an Excel file.
    df.to_excel(excel_writer, index_label='OTU', header=True)

    # Save the Excel file.
    excel_writer.close()


# Modify the column names of the first and second columns.
def rename():
    workbook = openpyxl.load_workbook('OTU_功能简写.xlsx')
    worksheet = workbook['Sheet1']
    worksheet.cell(row=1, column=1, value='OTU')
    worksheet.cell(row=1, column=2, value='function')
    workbook.save('OTU_功能简写.xlsx')


# Create a dictionary based on an Excel file.
def createdict(filename):
    df = pd.read_excel(filename)
    # Create an empty dictionary to store the converted data.
    data_dict = {}
    # Iterate over each row of the DataFrame, using the first column as the key and creating a list of the second and third columns as the values.
    for index, row in df.iterrows():
        key = row['key']
        value = [row['value1'], row['value2'], row['value3']]
        data_dict[key] = value
    return data_dict


# Remove duplicates and sort.
def sort_remove(inputs, order):
    sort = []
    for i in order:
        if i in inputs:
            sort.append(i)
        else:
            pass
    return sort


# Remove invalid otu
def remove_otu(df):
    rows_to_remove = []
    for index, row in df.iterrows():
        if row[tax.columns[0]] not in otuall:
            rows_to_remove.append(index)
    df = df.drop(rows_to_remove)
    return df


# Establish key-value pairs for OTU: functions
def tranfun(input):
    example = pd.read_excel(input)
    data_dict = {}
    for index, row in example.iterrows():
        data_dict[row['OTU']] = row['function']
    return data_dict


# Convert dataframe to excel
def createdfxl(filenames,outname):
    # Create DataFrame.
    df = pd.DataFrame(filenames)
    # Create Excel Writer object.
    excel_writer = pd.ExcelWriter(outname+'.xlsx', engine='openpyxl')
    # Write DataFrame to Excel file.
    df.to_excel(excel_writer, index=False)
    # Save and close excel
    excel_writer.close()





original_directory = input('请输入工作目录：')

# Get a list of all folders in the current directory.
folders = [f for f in os.listdir(original_directory) if os.path.isdir(os.path.join(original_directory, f))]
if not folders:
    txt_name = glob.glob((os.path.join(original_directory, "*.txt"))) # If my_list is empty, the condition is true.
    folders.append(txt_name[0])
for item in folders:
    os.chdir(original_directory)
    if len(folders) == 1:
        dealtxt(item)
    else:
        folder_path = os.path.join(original_directory, item)
        os.chdir(folder_path)  # Switch to a folder.
        txt_files = glob.glob("*.txt")
        txt = txt_files[0]  # Get the first txt file.
        dealtxt(txt)  # Preprocess the txt file and output it as 'deal.txt'

    files = open_file('deal.txt')  # Open the processed txt and save it to 'file'.

    # Retrieve all OTU using regular expressions and convert them into 'list'.
    otuall = list(set(re.findall(r'OTU_[0-9]+', files)))
    # Slice and filter different functionalities using "#" as a delimiter.
    otufun = files.split('#')
    del otufun[0]  # File processing, remove the first line with empty values.

    # For loop to remove unwanted OTUs.
    ziduan = pd.read_excel('字段替换.xlsx')
    allfuns = ziduan['key'].tolist()
    otufun_to_remove = []
    for remove in otufun:
        if not any(item in remove for item in allfuns):
            otufun_to_remove.append(remove)
    for remove in otufun_to_remove:
        otufun.remove(remove)

    # Create an empty dictionary to store 'OTU:function' key-value pairs.
    otufuns = {}
    otuws = {}
    # Create a dictionary for code replacements.
    my_dict = createdict('字段替换.xlsx')
    # For loop to filter OTU functions.
    order = ['C', 'H', 'O', 'N', 'S', 'Mn', 'Fe', 'As', 'OX', 'RX', 'FX']  # 功能顺序
    for o in otuall:
        key = o
        funf = []
        funw = []
        pattern = re.compile(r'\b' + re.escape(o) + r'\b')
        for otu in otufun:
            if pattern.search(otu):
                match = re.search(r'(\w+) \(\d+ records\)', otu)
                match = match.group(1)
                funw.append(match)
                # Replace 'functions' with corresponding codes.
                if match in my_dict:
                    for item in my_dict[match]:
                        if isinstance(item, str):
                            funf.append(item)
            else:
                pass

        remove_string = sort_remove(funf, order)  # Sort and remove duplicates and generating 'list'.
        result_string = ''.join(remove_string)  # Convert the 'list' to a 'string'.
        otufuns[key] = result_string  # Store the key-value pairs of 'OTU:function' in the 'dictionary otufuns'.
        otuws[key] = funw  # Store the key-value pairs of 'OTU:function' (complete) in the 'dictionary otuws'.

    createxl(otufuns,'OTU_功能简写')  # Convert the 'otufuns dictionary' to Excel.
    createxl(otuws, 'OTU_完整功能')  # Convert the 'otuws dictionary' to Excel.

    rename()  # Modify the names of the first two columns to "OTU" and "function".

    # Please convert .xls to .xlsx file format.
    taxname = glob.glob('*taxonomy*.xlsx')  # Use glob.glob() to find xlsx files with the "taxonomy" field.
    tax = pd.read_excel(taxname[0])
    tax = remove_otu(tax)  # Remove the unmatched OTUs.

    example = tranfun('OTU_功能简写.xlsx')  # Establish key-value pairs of OTU:function.

    tax['function'] = tax[tax.columns[0]].map(example)  # Match 'functions' with 'OTUs' and add them to the 'tax'.

    # Use the insert() method to insert the last column (function) into the second column position.
    last_column = tax.pop(tax.columns[-1])
    tax.insert(1, last_column.name, last_column)
    # Use the insert() method to insert the last column (taxonomy) into the third column
    index_of_taxonomy = tax.columns.get_loc('taxonomy')
    last_column = tax.pop(tax.columns[index_of_taxonomy])
    tax.insert(2, last_column.name, last_column)


    # write tax to excel
    writer = pd.ExcelWriter('report.xlsx')
    result_tax = tax.groupby('function', as_index=False).sum()  # Use function as the label classification and sum the number of OTUs
    result_tax = result_tax.drop(columns='taxonomy')  # Delete 'taxonomy' column
    result_tax.to_excel(writer, '汇总', index_label='NUM', header=True)  # Write to excel
    writer.close()  # close excel
    os.chdir(original_directory)

#Running successfully, please press any key to exit~
print("运行成功，请按任意键退出~")
ord(msvcrt.getch())
